"""
routers/export.py
  GET /api/export/dashboard  — filtered dashboard export (xlsx | csv)
  GET /api/export/calculator — two-section calculator export (xlsx | csv)
"""
import io
import re
from datetime import date
from typing import Optional

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse

from backend.database import get_connection

router = APIRouter()


def _slugify(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", text.lower()).strip("_")


# ── Dashboard Export ─────────────────────────────────────────────────────────

@router.get("/export/dashboard")
def export_dashboard(
    segment:  Optional[str] = Query(None),
    source:   Optional[str] = Query(None),
    category: Optional[str] = Query(None),
    q:        Optional[str] = Query(None),
    format:   str           = Query("xlsx"),
):
    with get_connection() as conn:
        sql = """
            SELECT
                i.name, i.id, i.segment, i.source, i.category,
                iv.period AS latest_period, iv.value AS latest_value,
                iv.mom_change, iv.yoy_change, i.base_year
            FROM indices i
            LEFT JOIN index_values iv
                ON iv.index_id = i.id
                AND iv.period = (SELECT MAX(period) FROM index_values WHERE index_id = i.id)
            WHERE i.active = 1
        """
        params = []
        if segment:
            sql += " AND i.segment = ?"
            params.append(segment)
        if source:
            sql += " AND i.source = ?"
            params.append(source)
        if category:
            sql += " AND i.category = ?"
            params.append(category)
        if q:
            sql += " AND LOWER(i.name) LIKE ?"
            params.append(f"%{q.lower()}%")
        sql += " ORDER BY i.segment, i.name"

        rows = conn.execute(sql, params).fetchall()

    if not rows:
        raise HTTPException(status_code=400, detail="No data to export for current filters")

    columns = ["Index Name", "Index ID", "Segment", "Source", "Category",
               "Latest Period", "Latest Value", "MoM %", "YoY %", "Base Year"]
    data = [
        [r["name"], r["id"], r["segment"], r["source"], r["category"],
         r["latest_period"], r["latest_value"],
         r["mom_change"], r["yoy_change"], r["base_year"]]
        for r in rows
    ]

    today = date.today().isoformat()
    if format == "csv":
        return _csv_response(columns, data, f"procurement_indices_{today}.csv")
    return _xlsx_dashboard_response(columns, data, f"procurement_indices_{today}.xlsx")


def _xlsx_dashboard_response(columns, data, filename):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Procurement Indices"

    # Header row — bold navy
    navy_fill = PatternFill("solid", fgColor="1A3A5C")
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = navy_fill
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"

    # Data rows
    for row_idx, row in enumerate(data, start=2):
        for col_idx, val in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _csv_response(columns, data, filename):
    import csv
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(columns)
    writer.writerows(data)
    csv_bytes = buf.getvalue().encode("utf-8-sig")
    return StreamingResponse(
        io.BytesIO(csv_bytes),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Calculator Export ────────────────────────────────────────────────────────

@router.get("/export/calculator")
def export_calculator(
    index_id:     str = Query(...),
    start_period: str = Query(...),
    end_period:   str = Query(...),
    format:       str = Query("xlsx"),
):
    with get_connection() as conn:
        meta = conn.execute(
            "SELECT name FROM indices WHERE id=?", (index_id,)
        ).fetchone()
        if not meta:
            raise HTTPException(status_code=404, detail=f"Index '{index_id}' not found")

        start_row = conn.execute(
            "SELECT value FROM index_values WHERE index_id=? AND period=?",
            (index_id, start_period),
        ).fetchone()
        end_row = conn.execute(
            "SELECT value FROM index_values WHERE index_id=? AND period=?",
            (index_id, end_period),
        ).fetchone()

        if not start_row:
            raise HTTPException(status_code=404, detail=f"No data for period {start_period}")
        if not end_row:
            raise HTTPException(status_code=404, detail=f"No data for period {end_period}")

        series = conn.execute(
            """SELECT period, value, mom_change, yoy_change
               FROM index_values
               WHERE index_id=? AND period >= ? AND period <= ?
               ORDER BY period ASC""",
            (index_id, start_period, end_period),
        ).fetchall()

    start_val  = start_row["value"]
    end_val    = end_row["value"]
    pct_change = round((end_val - start_val) / start_val * 100, 2) if start_val != 0 else 0.0
    abs_change = round(end_val - start_val, 4)
    index_name = meta["name"]

    slug = _slugify(index_name)
    filename = f"{slug}_{start_period}_to_{end_period}.{format}"

    if format == "csv":
        return _calc_csv_response(
            index_name, start_period, start_val, end_period, end_val,
            pct_change, abs_change, series, filename
        )
    return _calc_xlsx_response(
        index_name, start_period, start_val, end_period, end_val,
        pct_change, abs_change, series, filename
    )


def _calc_xlsx_response(index_name, start_period, start_val, end_period, end_val,
                         pct_change, abs_change, series, filename):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Calculator Result"

    # Section 1 — Summary card
    ws["A1"] = "Index"
    ws["B1"] = index_name
    ws["A2"] = "Start Period"
    ws["B2"] = start_period
    ws["A3"] = "Start Value"
    ws["B3"] = start_val
    ws["A4"] = "End Period"
    ws["B4"] = end_period
    ws["A5"] = "End Value"
    ws["B5"] = end_val
    ws["A6"] = "% Change"
    ws["B6"] = pct_change
    ws["A7"] = "Absolute Change"
    ws["B7"] = abs_change

    # Conditional formatting on % change cell
    pct_cell = ws["B6"]
    if pct_change >= 0:
        pct_cell.fill = PatternFill("solid", fgColor="C8E6C9")
    else:
        pct_cell.fill = PatternFill("solid", fgColor="FFCDD2")

    for row in range(1, 8):
        ws[f"A{row}"].font = Font(bold=True)

    # Spacer
    ws["A9"] = "Monthly Detail"
    ws["A9"].font = Font(bold=True, color="1A3A5C")

    # Section 2 — Monthly detail
    headers = ["Period", "Value", "MoM %", "YoY %"]
    navy_fill = PatternFill("solid", fgColor="1A3A5C")
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=10, column=col_idx, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = navy_fill

    for row_idx, r in enumerate(series, start=11):
        ws.cell(row=row_idx, column=1, value=r["period"])
        ws.cell(row=row_idx, column=2, value=r["value"])
        ws.cell(row=row_idx, column=3, value=r["mom_change"])
        ws.cell(row=row_idx, column=4, value=r["yoy_change"])

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 4, 12)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _calc_csv_response(index_name, start_period, start_val, end_period, end_val,
                        pct_change, abs_change, series, filename):
    import csv
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Summary"])
    w.writerow(["Index", index_name])
    w.writerow(["Start Period", start_period])
    w.writerow(["Start Value", start_val])
    w.writerow(["End Period", end_period])
    w.writerow(["End Value", end_val])
    w.writerow(["% Change", pct_change])
    w.writerow(["Absolute Change", abs_change])
    w.writerow([])
    w.writerow(["Monthly Detail"])
    w.writerow(["Period", "Value", "MoM %", "YoY %"])
    for r in series:
        w.writerow([r["period"], r["value"], r["mom_change"], r["yoy_change"]])

    csv_bytes = buf.getvalue().encode("utf-8-sig")
    return StreamingResponse(
        io.BytesIO(csv_bytes),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
