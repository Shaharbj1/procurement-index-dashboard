"""
routers/regional.py — GET /api/regional/summary  &  GET /api/export/regional
Serves the 52 regional indices (PPI/CPI/LCI/Energy × 13 countries).
EU Average is computed server-side as mean of the 11 core EU countries;
periods where fewer than 6 EU countries have data are excluded from the average.
"""
import io
import csv
from typing import Optional

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse

from backend.database import get_connection

router = APIRouter()

# ── Constants ─────────────────────────────────────────────────────────────────
_EU_COUNTRIES = [
    ("de", "Germany",     "DE"),
    ("fr", "France",      "FR"),
    ("it", "Italy",       "IT"),
    ("es", "Spain",       "ES"),
    ("nl", "Netherlands", "NL"),
    ("be", "Belgium",     "BE"),
    ("se", "Sweden",      "SE"),
    ("pl", "Poland",      "PL"),
    ("at", "Austria",     "AT"),
    ("pt", "Portugal",    "PT"),
    ("cz", "Czechia",     "CZ"),
]
_EU_ISO_LOWER = [iso for iso, _, _ in _EU_COUNTRIES]

_EXTRA_COUNTRIES = [
    ("ch", "Switzerland", "CH"),
    ("il", "Israel",      "IL"),
]

_ALL_COUNTRIES = _EU_COUNTRIES + _EXTRA_COUNTRIES   # 13 total

_TYPE_PERIOD_MAP = {
    "ppi":    "monthly",
    "cpi":    "monthly",
    "lci":    "quarterly",
    "energy": "semi-annual",
}

EU_AVG_MIN_COUNTRIES = 6   # minimum countries needed to compute EU average


def _index_id(type_: str, iso_lower: str) -> str:
    return f"reg_{type_}_{iso_lower}"


def _fetch_series(conn, index_id: str, n_periods: int) -> list[dict]:
    """Return the last n_periods rows for the given index, newest first then reversed."""
    rows = conn.execute(
        """SELECT period, value, mom_change, yoy_change
           FROM index_values
           WHERE index_id = ?
           ORDER BY period DESC
           LIMIT ?""",
        (index_id, n_periods),
    ).fetchall()
    rows.reverse()
    return [dict(r) for r in rows]


def _all_periods(series_map: dict[str, list[dict]]) -> list[str]:
    """Return sorted union of all period strings found across all country series."""
    periods = set()
    for rows in series_map.values():
        for r in rows:
            periods.add(r["period"])
    return sorted(periods)


def _build_value_map(rows: list[dict]) -> dict[str, float | None]:
    return {r["period"]: r["value"] for r in rows}


def _compute_eu_average(
    series_map: dict[str, list[dict]], periods: list[str]
) -> list[dict]:
    """
    Compute EU average across the 11 core EU countries per period.
    Periods with fewer than EU_AVG_MIN_COUNTRIES data points are set to None.
    """
    value_maps = {iso: _build_value_map(series_map.get(iso, [])) for iso in _EU_ISO_LOWER}
    result = []
    for p in periods:
        vals = [
            value_maps[iso][p]
            for iso in _EU_ISO_LOWER
            if p in value_maps[iso] and value_maps[iso][p] is not None
        ]
        avg = round(sum(vals) / len(vals), 4) if len(vals) >= EU_AVG_MIN_COUNTRIES else None
        result.append({"period": p, "value": avg, "mom_change": None, "yoy_change": None})
    return result


# ── Summary endpoint ───────────────────────────────────────────────────────────
@router.get("/summary")
def regional_summary(
    type: str = Query(..., regex="^(ppi|cpi|lci|energy)$"),
    periods: int = Query(24, ge=1, le=120),
):
    """
    Return the last `periods` data points for all 13 countries + EU average.
    Response shape:
      {
        "type":        "ppi",
        "period_type": "monthly",
        "periods":     ["2023-01", ...],
        "series": [
          {"country": "DE", "country_name": "Germany",
           "data": [{"period":"2023-01","value":102.1,...}, ...]},
          ...
          {"country": "EU_AVG", "country_name": "EU Average", "data": [...]}
        ]
      }
    """
    period_type = _TYPE_PERIOD_MAP[type]

    with get_connection() as conn:
        series_map: dict[str, list[dict]] = {}
        for iso_lower, _, _ in _ALL_COUNTRIES:
            idx_id = _index_id(type, iso_lower)
            series_map[iso_lower] = _fetch_series(conn, idx_id, periods)

    all_periods = _all_periods(series_map)
    eu_avg_data = _compute_eu_average(series_map, all_periods)

    series_out = []
    for iso_lower, country_name, iso_upper in _ALL_COUNTRIES:
        rows = series_map[iso_lower]
        vm = _build_value_map(rows)
        data = [
            {
                "period":     p,
                "value":      vm.get(p),
                "mom_change": next((r["mom_change"] for r in rows if r["period"] == p), None),
                "yoy_change": next((r["yoy_change"] for r in rows if r["period"] == p), None),
            }
            for p in all_periods
        ]
        series_out.append({
            "country":      iso_upper,
            "country_name": country_name,
            "data":         data,
        })

    # Append EU average last
    series_out.append({
        "country":      "EU_AVG",
        "country_name": "EU Average",
        "data":         eu_avg_data,
    })

    return {
        "type":        type,
        "period_type": period_type,
        "periods":     all_periods,
        "series":      series_out,
    }


# ── Export endpoint ────────────────────────────────────────────────────────────
@router.get("/export")
def regional_export(
    type: str = Query(..., regex="^(ppi|cpi|lci|energy)$"),
    periods: int = Query(24, ge=1, le=120),
    format: str = Query("xlsx", regex="^(xlsx|csv)$"),
):
    """Download regional data as Excel or CSV."""
    # Reuse summary logic
    payload = regional_summary(type=type, periods=periods)
    all_periods = payload["periods"]
    series = payload["series"]

    # Build a table: rows = periods, columns = countries
    header = ["Period"] + [s["country_name"] for s in series]
    rows_out = []
    for p in all_periods:
        row = [p]
        for s in series:
            dm = {d["period"]: d["value"] for d in s["data"]}
            val = dm.get(p)
            row.append("" if val is None else val)
        rows_out.append(row)

    type_label = type.upper()

    if format == "csv":
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(header)
        writer.writerows(rows_out)
        buf.seek(0)
        filename = f"regional_{type}_{periods}periods.csv"
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    # xlsx
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl not installed — cannot generate Excel file.",
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Regional {type_label}"

    # Header row styling
    hdr_font  = Font(bold=True, color="FFFFFF")
    hdr_fill  = PatternFill("solid", fgColor="1F4E79")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, h in enumerate(header, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font  = hdr_font
        cell.fill  = hdr_fill
        cell.alignment = hdr_align

    # Data rows — alternate shading
    alt_fill = PatternFill("solid", fgColor="EBF3FB")
    for r_idx, row_data in enumerate(rows_out, start=2):
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx % 2 == 0:
                cell.fill = alt_fill
            if c_idx > 1 and val != "":
                cell.number_format = "0.00"
                cell.alignment = Alignment(horizontal="right")

    # Column widths
    ws.column_dimensions["A"].width = 14
    for col_idx in range(2, len(header) + 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 16

    ws.freeze_panes = "B2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"regional_{type}_{periods}periods.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
